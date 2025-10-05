"""Tests for the top scorers Excel parser."""
from __future__ import annotations

import sys
from io import BytesIO
from types import ModuleType
from typing import Iterable, List

import pytest
from openpyxl import Workbook

from app.infrastructure.parsers.top_scorers_excel_parser import TopScorersExcelParser


def _build_workbook(rows: Iterable[Iterable[object]]) -> bytes:
    """Return workbook bytes built from ``rows`` in the active worksheet."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.delete_rows(1, worksheet.max_row)
    for row in rows:
        worksheet.append(list(row))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_top_scorers_parser_extracts_entries() -> None:
    """The parser should extract scorer entries and metadata from the spreadsheet."""

    rows: List[List[object]] = [
        ["LIGA AFICIONADOS F-11, 2ª AFICIONADOS F-11"],
        ["Temporada 2025-2026"],
        ["Jugador", "Equipo", "Grupo", "Partidos Jugados", "Goles", "Goles partido"],
        [
            "BOCANEGRA CAIPA, JOHN DAIRO",
            "CAFETERIA LA TACITA",
            "2ª AFICIONADOS F-11",
            3,
            "4",
            "1,3333",
        ],
        [
            "ARRIAGA MARTINEZ, MANUEL",
            "NEW COTTON MEKASO MCS",
            "2ª AFICIONADOS F-11",
            3,
            "5 (2 de penalti)",
            "1,6667",
        ],
    ]

    parser = TopScorersExcelParser()
    table = parser.parse(_build_workbook(rows))

    assert table.title == "LIGA AFICIONADOS F-11, 2ª AFICIONADOS F-11"
    assert table.competition == "LIGA AFICIONADOS F-11"
    assert table.category == "2ª AFICIONADOS F-11"
    assert table.season == "2025-2026"
    assert len(table.scorers) == 2

    first = table.scorers[0]
    second = table.scorers[1]

    assert first.player == "ARRIAGA MARTINEZ, MANUEL"
    assert first.team == "NEW COTTON MEKASO MCS"
    assert first.penalty_goals == 2
    assert first.goals_per_match == 1.6667

    assert second.player == "BOCANEGRA CAIPA, JOHN DAIRO"
    assert second.goals_total == 4
    assert second.goals_per_match == 1.3333


def test_top_scorers_parser_computes_ratio_when_missing() -> None:
    """The parser should derive goals per match when the column is empty."""

    rows = [
        ["LIGA AFICIONADOS F-11"],
        ["Temporada 2025-2026"],
        ["Jugador", "Equipo", "Grupo", "Partidos", "Goles", "Goles partido"],
        ["PLAYER ONE", "TEAM", "GRUPO", 2, "4", ""],
    ]

    parser = TopScorersExcelParser()
    table = parser.parse(_build_workbook(rows))

    scorer = table.scorers[0]
    assert scorer.matches_played == 2
    assert scorer.goals_total == 4
    assert scorer.goals_per_match == 2.0


def test_top_scorers_parser_accepts_decimal_separators() -> None:
    """The parser should accept ratios expressed with dots or commas."""

    rows = [
        ["LIGA AFICIONADOS F-11"],
        ["Temporada 2025-2026"],
        ["Jugador", "Equipo", "Grupo", "Partidos", "Goles", "Goles/Partido"],
        ["PLAYER ONE", "TEAM", "GRUPO", "2", "3", "1.5"],
        ["PLAYER TWO", "TEAM", "GRUPO", "3", "3", "1,0"],
    ]

    parser = TopScorersExcelParser()
    table = parser.parse(_build_workbook(rows))

    assert len(table.scorers) == 2
    assert table.scorers[0].goals_per_match == 1.5
    assert table.scorers[1].goals_per_match == 1.0


def test_top_scorers_parser_reads_html_xls_documents() -> None:
    """The parser should handle HTML tables saved with an XLS extension."""

    html = """
    <html>
        <body>
            <table>
                <tr><td colspan="6">LIGA AFICIONADOS F-11</td></tr>
                <tr><td colspan="6">Temporada 2025-2026</td></tr>
                <tr>
                    <th>Jugador</th>
                    <th>Equipo</th>
                    <th>Grupo</th>
                    <th>Partidos Jugados</th>
                    <th>Goles</th>
                    <th>Goles partido</th>
                </tr>
                <tr>
                    <td>PLAYER ONE</td>
                    <td>TEAM</td>
                    <td>GRUPO</td>
                    <td>2</td>
                    <td>3 (1 de penalti)</td>
                    <td>1,5</td>
                </tr>
            </table>
        </body>
    </html>
    """.strip()

    parser = TopScorersExcelParser()
    table = parser.parse(html.encode("utf-8"))

    assert table.title == "LIGA AFICIONADOS F-11"
    assert table.season == "2025-2026"
    assert len(table.scorers) == 1

    scorer = table.scorers[0]
    assert scorer.player == "PLAYER ONE"
    assert scorer.penalty_goals == 1
    assert scorer.goals_per_match == 1.5


def test_top_scorers_parser_preserves_trailing_cells_in_xls(monkeypatch) -> None:
    """The parser should keep trailing empty cells when using the XLS fallback."""

    from app.infrastructure.parsers import top_scorers_excel_parser as parser_module

    header = [
        "Jugador",
        "Equipo",
        "Grupo",
        "Partidos",
        "Goles",
        "Goles partido",
    ]
    data = ["PLAYER ONE", "TEAM", "GRUPO", 2, "4", ""]
    rows = [header, data]

    def fake_load_workbook(*_args: object, **_kwargs: object) -> None:
        raise ValueError("legacy xls")

    def fake_loader(document_bytes: bytes) -> List[List[object]]:
        assert document_bytes == b"fake-xls"
        return [list(row) for row in rows]

    monkeypatch.setattr(parser_module, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(parser_module, "_XLS_LOADERS", [fake_loader])

    parser = parser_module.TopScorersExcelParser()
    table = parser.parse(b"fake-xls")

    assert len(table.scorers) == 1
    scorer = table.scorers[0]
    assert scorer.matches_played == 2
    assert scorer.goals_total == 4
    assert scorer.goals_per_match == 2.0


def test_build_xls_loaders_prefers_supported_modules(monkeypatch) -> None:
    """The loader discovery should skip unsupported xlrd releases."""

    from app.infrastructure.parsers import top_scorers_excel_parser as parser_module

    unsupported = ModuleType("xlrd")
    unsupported.__version__ = "2.0.1"
    unsupported.open_workbook = lambda **_kwargs: None  # type: ignore[attr-defined]

    class DummySheet:
        nrows = 1
        ncols = 1

        @staticmethod
        def row_values(_index: int, end_colx: int | None = None) -> List[object]:
            return ["value"] if end_colx else ["value"]

    class DummyBook:
        @staticmethod
        def sheet_by_index(index: int) -> DummySheet:
            assert index == 0
            return DummySheet()

    supported = ModuleType("xlrd3")
    supported.__version__ = "1.1.0"
    supported.open_workbook = lambda **_kwargs: DummyBook()  # type: ignore[attr-defined]

    modules = {"xlrd": unsupported, "xlrd3": supported}

    def fake_import_module(name: str) -> ModuleType:
        if name not in modules:
            raise ModuleNotFoundError(name)
        return modules[name]

    monkeypatch.setattr(parser_module, "import_module", fake_import_module)

    loaders = parser_module._build_xls_loaders()

    assert len(loaders) == 2
    assert loaders[0](b"fake") == [["value"]]
    with pytest.raises(ValueError):
        loaders[1](b"fake")


def test_build_xls_loaders_uses_pyexcel_when_available(monkeypatch) -> None:
    """The loader discovery should use pyexcel-xls when installed."""

    from app.infrastructure.parsers import top_scorers_excel_parser as parser_module

    pyexcel_module = ModuleType("pyexcel_xls")

    def fake_get_data(
        stream: BytesIO, *, file_type: str | None = None
    ) -> dict[str, List[List[object]]]:
        assert isinstance(stream, BytesIO)
        assert file_type == "xls"
        return {"Hoja1": [["Jugador", "Equipo", "Grupo"], ["NAME", "TEAM", "GROUP"]]}

    pyexcel_module.get_data = fake_get_data  # type: ignore[attr-defined]

    modules = {"pyexcel_xls": pyexcel_module}

    def fake_import_module(name: str) -> ModuleType:
        if name not in modules:
            raise ModuleNotFoundError(name)
        return modules[name]

    monkeypatch.setattr(parser_module, "import_module", fake_import_module)

    loaders = parser_module._build_xls_loaders()

    assert len(loaders) == 2
    assert loaders[0](b"fake") == [
        ["Jugador", "Equipo", "Grupo"],
        ["NAME", "TEAM", "GROUP"],
    ]
    with pytest.raises(ValueError):
        loaders[1](b"fake")
