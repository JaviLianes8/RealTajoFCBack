"""Excel parser for extracting top scorers tables."""
from __future__ import annotations

import re
from html import unescape
from html.parser import HTMLParser
from importlib import import_module
from io import BytesIO
from typing import Any, Callable, Dict, Iterable, List, Optional

from openpyxl import load_workbook

def _parse_version_tuple(raw_version: str) -> tuple[int, ...]:
    """Return a tuple with the numeric portions of ``raw_version``."""

    parts = re.findall(r"\d+", raw_version)
    return tuple(int(part) for part in parts)


def _build_xls_loaders() -> List[Callable[[bytes], List[List[Any]]]]:
    """Return XLS loader callables using the available optional dependencies."""

    loaders: List[Callable[[bytes], List[List[Any]]]] = []

    for module_name in ("xlrd", "xlrd3"):
        try:  # pragma: no cover - depends on optional runtime dependencies
            module = import_module(module_name)
        except Exception:  # pragma: no cover - module not installed or unusable
            continue

        version = _parse_version_tuple(getattr(module, "__version__", ""))
        if module_name == "xlrd" and version >= (2, 0, 0):
            # xlrd >= 2 removed support for legacy XLS files, fall back to xlrd3.
            continue

        open_workbook = getattr(module, "open_workbook", None)
        if not callable(open_workbook):
            continue

        def _load_with_xlrd(
            document_bytes: bytes, *, _module: Any = module
        ) -> List[List[Any]]:
            book = _module.open_workbook(file_contents=document_bytes)
            sheet = book.sheet_by_index(0)
            total_columns = getattr(sheet, "ncols", 0)
            rows: List[List[Any]] = []
            for index in range(sheet.nrows):
                if total_columns:
                    values = list(sheet.row_values(index, end_colx=total_columns))
                    if len(values) < total_columns:
                        values.extend([""] * (total_columns - len(values)))
                else:
                    values = list(sheet.row_values(index))
                rows.append(values)
            return rows

        loaders.append(_load_with_xlrd)

    try:  # pragma: no cover - depends on optional runtime dependencies
        pyexcel_module = import_module("pyexcel_xls")
    except Exception:  # pragma: no cover - module not installed or unusable
        pyexcel_module = None

    if pyexcel_module is not None:
        get_data = getattr(pyexcel_module, "get_data", None)
        if callable(get_data):

            def _load_with_pyexcel(document_bytes: bytes) -> List[List[Any]]:
                data = get_data(BytesIO(document_bytes), file_type="xls")
                if not data:
                    return []
                first_sheet = next(iter(data.values()), [])
                max_length = max((len(row) for row in first_sheet), default=0)
                normalised_rows: List[List[Any]] = []
                for row in first_sheet:
                    values = list(row)
                    if max_length and len(values) < max_length:
                        values.extend([""] * (max_length - len(values)))
                    normalised_rows.append(values)
                return normalised_rows

            loaders.append(_load_with_pyexcel)

    def _load_with_html(document_bytes: bytes) -> List[List[Any]]:
        """Return rows extracted from HTML tables masquerading as XLS files."""

        text = _decode_html_document(document_bytes)
        if text is None:
            raise ValueError("The provided document is not an HTML table.")

        parser = _HTMLTableParser()
        parser.feed(text)
        parser.close()
        if not parser.rows:
            raise ValueError("No tables were found in the HTML document.")

        max_length = max((len(row) for row in parser.rows), default=0)
        normalised: List[List[Any]] = []
        for row in parser.rows:
            values = list(row)
            if max_length and len(values) < max_length:
                values.extend([""] * (max_length - len(values)))
            normalised.append(values)
        return normalised

    loaders.append(_load_with_html)

    return loaders


_XLS_LOADERS = _build_xls_loaders()

from app.application.process_top_scorers import TopScorersParser
from app.domain.models.top_scorers import TopScorerEntry, TopScorersTable


_HEADER_ALIASES: Dict[str, set[str]] = {
    "player": {"jugador"},
    "team": {"equipo"},
    "group": {"grupo"},
    "matches": {"partidos", "partidos jugados"},
    "goals": {"goles"},
    "ratio": {"goles partido", "goles/partido", "goles por partido"},
}
_PENALTIES_RE = re.compile(r"(\d+)\s*de\s*penalti", re.IGNORECASE)
_NUMBER_RE = re.compile(r"\d+")


def _decode_html_document(document_bytes: bytes) -> Optional[str]:
    """Return the decoded HTML string when ``document_bytes`` contains a table."""

    candidates = (
        "utf-8",
        "utf-8-sig",
        "utf-16",
        "utf-16le",
        "utf-16be",
        "latin-1",
        "windows-1252",
    )
    marker = "<table"

    for encoding in candidates:
        try:
            text = document_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
        lower_text = text.lower()
        if marker in lower_text:
            return text
    return None


class _HTMLTableParser(HTMLParser):
    """Collect cell values from the first HTML table found in a document."""

    def __init__(self) -> None:
        super().__init__()
        self._inside_table = False
        self._nested_tables = 0
        self._current_row: Optional[List[str]] = None
        self._current_cell: Optional[List[str]] = None
        self._current_col = 0
        self._pending_rowspans: Dict[int, tuple[int, str]] = {}
        self._cell_colspan = 1
        self._cell_rowspan = 1
        self.rows: List[List[str]] = []

    # ------------------------------------------------------------------
    # HTMLParser API
    # ------------------------------------------------------------------
    def handle_starttag(self, tag: str, attrs: List[tuple[str, Optional[str]]]) -> None:
        if tag == "table":
            if self._inside_table:
                self._nested_tables += 1
            else:
                self._inside_table = True
        elif not self._inside_table:
            return
        elif tag == "tr":
            self._start_row()
        elif tag in {"td", "th"}:
            self._start_cell(dict(attrs))
        elif tag == "br" and self._current_cell is not None:
            self._current_cell.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag == "table" and self._inside_table:
            if self._nested_tables:
                self._nested_tables -= 1
            else:
                self._inside_table = False
                self._pending_rowspans.clear()
        elif not self._inside_table:
            return
        elif tag == "tr":
            self._end_row()
        elif tag in {"td", "th"}:
            self._end_cell()

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_entityref(self, name: str) -> None:  # pragma: no cover - inherited API
        if self._current_cell is not None:
            self._current_cell.append(unescape(f"&{name};"))

    def handle_charref(self, name: str) -> None:  # pragma: no cover - inherited API
        if self._current_cell is not None:
            self._current_cell.append(unescape(f"&#{name};"))

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _start_row(self) -> None:
        self._current_row = []
        self._current_col = 0
        self._consume_pending_rowspans()

    def _end_row(self) -> None:
        if self._current_row is None:
            return
        self._consume_pending_rowspans()
        self.rows.append(self._current_row)
        self._current_row = None
        self._current_col = 0

    def _start_cell(self, attrs: Dict[str, Optional[str]]) -> None:
        if self._current_row is None:
            self._start_row()
        self._consume_pending_rowspans()
        self._current_cell = []
        self._cell_colspan = self._parse_span(attrs.get("colspan"))
        self._cell_rowspan = self._parse_span(attrs.get("rowspan"))

    def _end_cell(self) -> None:
        if self._current_row is None or self._current_cell is None:
            return
        value = unescape("".join(self._current_cell)).replace("\xa0", " ").strip()
        for offset in range(self._cell_colspan):
            column_index = self._current_col + offset
            column_value = value if offset == 0 else ""
            self._current_row.append(column_value)
            if self._cell_rowspan > 1:
                self._pending_rowspans[column_index] = (
                    self._cell_rowspan - 1,
                    column_value,
                )
        self._current_col += self._cell_colspan
        self._current_cell = None
        self._cell_colspan = 1
        self._cell_rowspan = 1

    def _consume_pending_rowspans(self) -> None:
        while self._pending_rowspans.get(self._current_col):
            remaining, value = self._pending_rowspans[self._current_col]
            self._current_row.append(value)
            if remaining > 1:
                self._pending_rowspans[self._current_col] = (remaining - 1, value)
            else:
                del self._pending_rowspans[self._current_col]
            self._current_col += 1

    @staticmethod
    def _parse_span(raw: Optional[str]) -> int:
        try:
            span = int(raw) if raw else 1
        except (TypeError, ValueError):
            span = 1
        return max(span, 1)



def _stringify(value: Any) -> str:
    """Return a trimmed string representation for ``value`` suitable for JSON."""

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    """Return a normalised lowercase header representation."""

    return " ".join(_stringify(value).lower().split())


def _row_is_empty(row: Iterable[Any]) -> bool:
    """Return ``True`` when every cell in ``row`` is blank."""

    return all(not _stringify(cell) for cell in row)


def _load_excel_rows(document_bytes: bytes) -> List[List[Any]]:
    """Load spreadsheet rows using ``openpyxl`` with optional XLS fallbacks."""

    stream = BytesIO(document_bytes)
    try:
        workbook = load_workbook(stream, data_only=True, read_only=True)
    except Exception as openpyxl_error:  # pragma: no cover - exercised in XLS fallback
        errors: List[Exception] = []
        for loader in _XLS_LOADERS:
            try:
                return loader(document_bytes)
            except Exception as loader_error:  # pragma: no cover - defensive fallback path
                errors.append(loader_error)
        root_error = errors[-1] if errors else openpyxl_error
        raise ValueError("The provided Excel file could not be parsed.") from root_error
    else:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _locate_header(rows: List[List[Any]]) -> tuple[int, Dict[str, int]]:
    """Return the header row index plus a mapping between logical and column indices."""

    for index, row in enumerate(rows):
        if _row_is_empty(row):
            continue
        normalised = [_normalize_header(cell) for cell in row]
        mapping: Dict[str, int] = {}
        for key, aliases in _HEADER_ALIASES.items():
            for col_index, cell in enumerate(normalised):
                if cell in aliases:
                    mapping[key] = col_index
                    break
        if len(mapping) == len(_HEADER_ALIASES):
            return index, mapping
    raise ValueError("The Excel sheet does not contain the expected scorer headers.")


def _extract_metadata(rows: Iterable[List[Any]]) -> tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """Derive table metadata (title, competition, category, season) from ``rows``."""

    lines: List[str] = []
    for row in rows:
        text = " ".join(filter(None, (_stringify(cell) for cell in row))).strip()
        if text:
            lines.append(text)

    if not lines:
        return None, None, None, None

    title = lines[0]
    season_index = next((idx for idx, line in enumerate(lines) if "Temporada" in line), None)

    season: Optional[str] = None
    descriptor: Optional[str] = None
    if season_index is not None:
        season_line = lines[season_index]
        before, _, after = season_line.partition("Temporada")
        season = _stringify(after) or None
        descriptor = _stringify(before).strip(",") or None
        if not descriptor and season_index > 0:
            descriptor = lines[season_index - 1]
    else:
        descriptor = title

    competition: Optional[str] = None
    category: Optional[str] = None
    if descriptor:
        if "," in descriptor:
            first, _, rest = descriptor.partition(",")
            competition = first.strip() or None
            category = rest.strip() or None
        else:
            competition = descriptor.strip() or None

    return title or None, competition, category, season


def _parse_int_cell(value: Any) -> Optional[int]:
    """Return an integer extracted from ``value`` when possible."""

    text = _stringify(value)
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        try:
            return int(float(text.replace(",", ".")))
        except ValueError:
            return None


def _parse_ratio_cell(value: Any) -> Optional[float]:
    """Return a float extracted from ``value`` when possible."""

    text = _stringify(value)
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _parse_goals_cell(value: Any) -> tuple[Optional[int], Optional[int], Optional[str]]:
    """Return total goals, penalty goals and the textual detail from ``value``."""

    text = _stringify(value)
    if not text:
        return None, None, None

    match = _PENALTIES_RE.search(text)
    penalties = int(match.group(1)) if match else None
    numbers = _NUMBER_RE.findall(text)
    total = int(numbers[0]) if numbers else None
    details = text if text else (str(total) if total is not None else None)
    return total, penalties, details


class TopScorersExcelParser(TopScorersParser):
    """Decode top scorers information from uploaded Excel spreadsheets."""

    def parse(self, document_bytes: bytes) -> TopScorersTable:
        """Parse ``document_bytes`` into a :class:`TopScorersTable` instance."""

        rows = _load_excel_rows(document_bytes)
        if not rows:
            raise ValueError("The Excel sheet does not contain any data.")

        header_index, mapping = _locate_header(rows)
        title, competition, category, season = _extract_metadata(rows[:header_index])

        scorers: List[TopScorerEntry] = []
        for raw_row in rows[header_index + 1 :]:
            if _row_is_empty(raw_row):
                continue

            player = _stringify(raw_row[mapping["player"]])
            if not player:
                continue

            team = _stringify(raw_row[mapping["team"]]) or None
            group = _stringify(raw_row[mapping["group"]]) or None
            matches = _parse_int_cell(raw_row[mapping["matches"]])
            goals_total, penalty_goals, goals_details = _parse_goals_cell(
                raw_row[mapping["goals"]]
            )
            ratio = _parse_ratio_cell(raw_row[mapping["ratio"]])
            if ratio is None and matches and goals_total is not None and matches != 0:
                ratio = goals_total / matches

            raw_lines = [
                value
                for value in (
                    _stringify(raw_row[mapping["player"]]),
                    _stringify(raw_row[mapping["team"]]),
                    _stringify(raw_row[mapping["group"]]),
                    _stringify(raw_row[mapping["matches"]]),
                    _stringify(raw_row[mapping["goals"]]),
                    _stringify(raw_row[mapping["ratio"]]),
                )
                if value
            ]

            scorers.append(
                TopScorerEntry(
                    player=player,
                    team=team,
                    group=group or category,
                    matches_played=matches,
                    goals_total=goals_total,
                    goals_details=goals_details,
                    penalty_goals=penalty_goals,
                    goals_per_match=ratio,
                    raw_lines=raw_lines,
                )
            )

        if not scorers:
            raise ValueError("No scorer entries were found in the provided Excel file.")

        indexed = list(enumerate(scorers))
        indexed.sort(key=lambda item: (-(item[1].goals_total or -1), item[0]))
        ordered = [entry for _, entry in indexed]

        return TopScorersTable(
            title=title,
            competition=competition,
            category=category,
            season=season,
            scorers=ordered,
        )
